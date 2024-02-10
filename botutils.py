# Bot functions
# Version 3.1
# Ben Samans, Updated 12/24/2023

import interactions
from emoji import emojize, demojize
import json
import botdata as bd
from time import strftime
from termcolor import colored
from colorama import init
import yaml
import asyncio
from os import path, mkdir, remove
from random import randint, shuffle
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, borders
from excel2img import export_img
from random import choice
from typing import Union
from datetime import datetime, timedelta
import matplotlib.pyplot as plt
import io
import matplotlib.font_manager
from PIL import Image
from PIL import ImageFont
from PIL import ImageDraw
from PIL import ImageColor
init()

# Class Definitions


class Response:
    def __init__(self, exact, trig, text, user_id):
        self.exact = exact
        self.trig = trig
        self.text = text
        self.user_id = user_id

    def __repr__(self):
        return f'<Exact={self.exact}> <Trig={self.trig}> <Text={self.text}> <User_id={self.user_id}>'

    def asdict(self):
        return {'exact': self.exact, 'trig': self.trig, 'text': self.text, 'user_id': self.user_id}

    def add_rsp_text(self, new_text):
        if not isinstance(self.text, list):
            self.text = list(self.text)
        self.text.append(new_text)


class ListMsg:
    def __init__(
            self, num: int, page: int, guild: interactions.Guild, channel: interactions.BaseChannel,
            msg_type: str, payload=None
    ):
        self.num = num
        self.page = page
        self.guild = guild
        self.channel = channel
        self.msg_type = msg_type
        self.payload = payload


class TrainShot:
    def __init__(self, location: tuple, genre: str, info: str, time: str):
        self.location = location
        self.genre = genre
        self.info = info
        self.time = time

    def asdict(self):
        return {"location": self.location, "genre": self.genre, "info": self.info, "time": self.time}

    def __repr__(self):
        return f"<location={self.location}> <genre={self.genre}> <info={self.info}> <time={self.time}"

    def to_json(self):
        return json.dumps(self, default=lambda o: o.__dict__, sort_keys=True, indent=4)


class TrainPlayer:
    def __init__(
            self, member: interactions.Member = None, tag: str = None, dmchannel: interactions.DMChannel = None,
            rails: int = 0, shots: list[TrainShot] = None, vis_tiles: list[tuple] = None, score: dict = None,
            start: tuple = None, end: tuple = None, done: bool = False, donetime: str = None
    ):
        if vis_tiles is None:
            vis_tiles = []
        if score is None:
            score = {}
        if shots is None:
            shots = []
        self.member = member
        self.tag = tag
        self.done = done
        self.rails = rails
        self.dmchannel = dmchannel
        self.start = start
        self.end = end
        self.score = score
        self.shots = shots
        self.donetime = donetime
        self.vis_tiles = vis_tiles

    def __repr__(self):
        return f"<member={self.member}> <tag={self.tag}> <done={self.done}> <rails={self.rails}> " \
               f"<dmchannel={self.dmchannel}> <start={self.start}> <end={self.end}> <score={self.score}> " \
               f"<shots={self.shots}> <donetime={self.donetime}> <vis_tiles={self.vis_tiles}>"

    def asdict(self):
        shot_list = []
        for shot in self.shots:
            shot_list.append(shot.asdict())
        return {
            "member_id": self.member.id, "tag": self.tag, "done": self.done, "rails": self.rails,
            "dmchannel": self.dmchannel.id, "start": self.start, "end": self.end, "score": self.score,
            "shots": shot_list, "donetime": self.donetime, "vis_tiles": self.vis_tiles
        }

    def get_shot_genre_count(self) -> dict:
        shot_dict = {}
        for shot in self.shots:
            if shot.genre in shot_dict:
                shot_dict[shot.genre] += 1
            else:
                shot_dict[shot.genre] = 1
        return shot_dict


class TrainGame:
    def __init__(
            self, name: str = None, date: str = None, players: list[TrainPlayer] = None,
            board: dict = None, gameid: int = None, active: bool = True, size: tuple[int, int] = None
    ):
        self.name = name
        self.date = date
        self.players = players
        self.gameid = gameid
        self.active = active
        self.size = size
        self.board = board

    def asdict(self) -> dict:
        player_list = []
        for player in self.players:
            player_list.append(player.asdict())
        board_dict = {}
        for key in self.board.keys():
            board_dict[str(key)] = self.board[key]
        return {
            "name": self.name, "date": self.date, "players": player_list,
            "gameid": self.gameid, "active": self.active, "size": self.size, "board": board_dict,
        }

    def __repr__(self):
        return f"<name={self.name}> <date={self.date}> <players={self.players}> <gameid={self.gameid}> " \
               f"<active={self.active}> <size={self.size}> <board={self.board}> "

    def is_done(self) -> bool:
        done = True
        for player in self.players:
            if not player.done:
                done = False
                break
        return done

    def get_player(self, player_id: int) -> Union[tuple[None, None], tuple[int, TrainPlayer]]:
        player = None
        player_idx = None
        for idx, p in enumerate(self.players):
            if p.member.id == player_id:
                player = p
                player_idx = idx
        return player_idx, player

    def save_game(self, filepath: str) -> None:
        with open(f"{filepath}/gamedata.json", "w") as f:
            json.dump(self.asdict(), f, indent=4)

    def gen_trains_board(
            self, board_size: tuple[int, int] = (16, 16), river_ring: int = 0
    ):
        width = board_size[0] + 2*river_ring
        height = board_size[1] + 2*river_ring

        def next_to_resource(tilepos, resource):
            # Returns true if the grid tile is directly adjacent to the specified resource
            x = tilepos[0]
            y = tilepos[1]

            for x_new in (x-1, x+1):
                try:
                    if self.board[(x_new, y)]["resource"] == resource:
                        return True
                except KeyError:
                    pass

            for y_new in (y-1, y+1):
                try:
                    if self.board[(x, y_new)]["resource"] == resource:
                        return True
                except KeyError:
                    pass

            return False

        def near_resource(tilepos, resource, spread):
            # Returns true if the grid tile is within {spread} tiles of the specified resource
            # Tilepos is (x, y)
            x = tilepos[0]
            y = tilepos[1]

            for x_new in range(x-spread, x+spread):
                for y_new in range(y-spread, y+spread):
                    try:
                        if self.board[(x_new, y_new)]["resource"] == resource:
                            return True
                    except KeyError:
                        pass

            return False

        def generate_random_resources(tilepos):
            # Tilepos is (x, y)
            # Determines based on the chances below (out of 1000) if a tile will be populated with a resource.
            # Only adds to empty tiles (after count based resources have been placed)

            wheat_chance = 100
            wheat_near_chance = 50

            # gem_chance = 10
            # gem_near_chance = 1

            # prison_chance = 8

            wood_chance = 130
            wood_near_chance = 30

            # city_chance = 16

            house_chance = 70
            house_near_chance = 35

            if self.board[tilepos]["resource"] is not None:
                return self.board[tilepos]["resource"]

            if self.board[tilepos]["terrain"] is not None:
                return None

            # Wheat
            if near_resource(tilepos, '🌾', 3):
                if randint(1, 1000) <= wheat_near_chance:
                    return '🌾'
            else:
                if randint(1, 1000) <= wheat_chance:
                    return '🌾'
            '''
            # Gems
            if near_resource(self.board, tilepos, '💎'):
                if randint(1, 1000) <= gem_near_chance:
                    return '💎'
            else:
                if randint(1, 1000) <= gem_chance:
                    return '💎'
            '''
            # Wood
            if near_resource(tilepos, '🌳', 2):
                if randint(1, 1000) <= wood_near_chance:
                    return '🌳'
            else:
                if randint(1, 1000) <= wood_chance:
                    return '🌳'
            '''
            # Cities
            if randint(1, 1000) <= city_chance:
                return '🌃'
            '''
            # Houses
            if next_to_resource(tilepos, '🏠'):
                if randint(1, 1000) <= house_near_chance:
                    return '🏠'
            else:
                if randint(1, 1000) <= house_chance:
                    return '🏠'
            '''
            # Prisons
            if randint(1, 1000) <= prison_chance:
                return '⛓'
            '''
            return None

        def generate_count_resource(count, resource, min_spread=0):
            # Grid Size is (x, y)
            added = 0
            attempts = 0
            while added < count:
                attempts += 1

                y = randint(1, width)
                x = randint(1, width)

                # Add resource if tile is empty and meets the minimum spread requirement
                if self.board[(x, y)]["resource"] is None and self.board[(x, y)]["terrain"] is None \
                        and not near_resource((x, y), resource, min_spread):
                    self.board[(x, y)]["resource"] = resource
                    added += 1
                    attempts = 0
                # Reduces minimum spread requirement at 15 attempts and tries again
                if attempts > 15:
                    min_spread -= 1
                    attempts = 0
                # Infinite loop catch if unable to add resource
                if min_spread <= 0:
                    break
            return self.board

        def generate_zones():
            # Adds genre zones in randomized order to grid. Both dimensions of the grid must be divisible by 4 to allow
            # for 16 zones.

            def add_zone(z_width, z_height, start_pos, genre):
                for row in range(start_pos[0], start_pos[0] + z_height):
                    for col in range(start_pos[1], start_pos[1] + z_width):
                        self.board[(row, col)]["zone"] = genre

            if (width - 2*river_ring) % 4 != 0 or (height - 2*river_ring) % 4 != 0:
                return self.board

            zone_width = (width - 2*river_ring)//4
            zone_height = (height - 2*river_ring)//4
            zone_order = list(genre_colors.keys())
            shuffle(zone_order)

            for i in range(16):
                zone_pos = (
                    zone_height*(i // 4) + river_ring + 1,
                    zone_width*(i % 4) + river_ring + 1
                )
                add_zone(zone_width, zone_height, zone_pos, zone_order[i])
            return self.board

        def generate_river(direction):

            base_chance = 1
            avg_width = 1.5
            density = 2.2

            river_tiles = []

            if direction == 'Right':
                river_start = (
                    randint(round(width*0.25), round(width*0.75)),
                    randint(1, round(width*0.25))
                )
                river_tiles.append(river_start)
                river_center = [river_start[0]]

                for col in range(river_start[1] + 1, width + 1):
                    if randint(1, 1000) <= base_chance*1000:
                        river_center.append(river_center[-1] + randint(-1, 1))
                    else:
                        break
                    for row in range(1, height + 1):
                        diff = abs(row - river_center[-1])
                        chance = 1000*(base_chance - 0.25/avg_width*diff**density)
                        if randint(1, 1000) <= chance:
                            river_tiles.append((row, col))

            elif direction == 'DownRight':
                river_start = (
                    randint(1, round(width*0.25)),
                    randint(1, round(width*0.25))
                )
                river_tiles.append(river_start)
                river_center = [river_start[0]]

                for col in range(river_start[1] + 1, width + 1):
                    if randint(1, 1000) <= base_chance*1000:
                        river_center.append(river_center[-1] + randint(0, 2))
                    else:
                        break
                    for row in range(1, height + 1):
                        diff = abs(row - river_center[-1])
                        chance = 1000*(base_chance - 0.25/avg_width*diff**density)
                        if randint(1, 1000) <= chance:
                            river_tiles.append((row, col))

            elif direction == 'Down':
                river_start = (
                    randint(1, round(width*0.25)),
                    randint(round(width*0.25), round(width*0.75))
                )
                river_tiles.append(river_start)
                river_center = [river_start[1]]

                for row in range(river_start[0] + 1, height + 1):
                    if randint(1, 1000) <= base_chance*1000:
                        river_center.append(river_center[-1] + randint(-1, 1))
                    else:
                        break
                    for col in range(1, width + 1):
                        diff = abs(col - river_center[-1])
                        chance = 1000*(base_chance - 0.25/avg_width*diff**density)
                        if randint(1, 1000) <= chance:
                            river_tiles.append((row, col))

            elif direction == 'DownLeft':
                river_start = (
                    randint(round(height*0.75), height),
                    randint(1, round(width*0.25))
                )
                river_tiles.append(river_start)
                river_center = [river_start[0]]

                for col in range(river_start[1] + 1, width + 1):
                    if randint(1, 1000) <= base_chance*1000:
                        river_center.append(river_center[-1] + randint(-2, 0))
                    else:
                        break
                    for row in range(1, height + 1):
                        diff = abs(row - river_center[-1])
                        chance = 1000*(base_chance - 0.25/avg_width*diff**density)
                        if randint(1, 1000) <= chance:
                            river_tiles.append((row, col))

            else:
                return self.board

            for pos in river_tiles:
                self.board[pos]["terrain"] = "river"
            return self.board

        # For zones to generate, both board dimensions must be divisible by 4

        # Generate empty board
        self.board = {}
        for r in range(height):
            for c in range(width):
                if r+1 <= river_ring or r+1 > height - river_ring:
                    terrain = "river"
                elif c+1 <= river_ring or c+1 > width - river_ring:
                    terrain = "river"
                else:
                    terrain = None
                self.board[(r+1, c+1)] = {"resource": None, "terrain": terrain, "zone": None, "rails": []}

        # Add terrain (chance out of 1000)
        river_chance = 900
        if randint(1, 1000) <= river_chance:
            river_dir = ('Down', 'DownLeft', 'DownRight', 'Right')
            self.board = generate_river(choice(river_dir))

        # Add count-based resources
        city_count = 4
        prison_count = 2
        gem_count = 2

        self.board = generate_count_resource(city_count, '🌃', 4)
        self.board = generate_count_resource(prison_count, '⛓', 6)
        self.board = generate_count_resource(gem_count, '💎', 8)

        # Add random resources
        for c in range(width):
            for r in range(height):
                self.board[(r+1, c+1)]["resource"] = generate_random_resources((r+1, c+1))

        # Add genre zones
        self.board = generate_zones()
        return None

    def add_visible_tiles(self, player_idx,  shot_row: int, shot_col: int):
        width = self.size[0]
        height = self.size[1]
        dist = 2
        for row in range(shot_row - dist, shot_row + dist + 1):
            for col in range(shot_col - dist, shot_col + dist + 1):
                if (row, col) in self.players[player_idx].vis_tiles:  # Already rendered tiles
                    pass
                elif 1 <= row <= height and 1 <= col <= width:  # Out of bounds tiles
                    self.players[player_idx].vis_tiles.append((row, col))

    def gen_player_locations(self, river_ring: int) -> Union[None, bool]:
        row_bounds = (1+river_ring, self.size[1]-river_ring)
        col_bounds = (1+river_ring, self.size[0]-river_ring)
        for player_idx, player in enumerate(self.players):
            quadrant = choice(("Left", "Top"))
            taken_spaces = []

            # Generate start locations (NOTE: game.size is (width, height) while coordinates are in (row, col)
            attempts = 0
            start_loc = None
            while attempts <= 40:
                if quadrant == "Left":
                    start_loc = (randint(row_bounds[0], row_bounds[1]), col_bounds[0])
                else:
                    start_loc = (row_bounds[0], randint(col_bounds[0], col_bounds[1]))

                if self.board[start_loc]["terrain"] is None \
                        and self.board[start_loc]["resource"] is None \
                        and start_loc not in taken_spaces:
                    player.start = start_loc
                    taken_spaces.append(start_loc)
                    break
                attempts += 1
            if start_loc is None or attempts > 40:  # Error catch
                return True
            self.add_visible_tiles(player_idx, start_loc[0], start_loc[1])

            # Generate end locations
            attempts = 0
            end_loc = None
            while attempts <= 40:
                if quadrant == "Left":
                    end_loc = (randint(row_bounds[0], row_bounds[1]), col_bounds[1])
                else:
                    end_loc = (row_bounds[1], randint(col_bounds[0], col_bounds[1]))

                if self.board[end_loc]["terrain"] is None \
                        and self.board[end_loc]["resource"] is None \
                        and end_loc not in taken_spaces:
                    player.end = end_loc
                    taken_spaces.append(end_loc)
                    break
                attempts += 1
            if end_loc is None or attempts > 40:  # Error catch
                return True
            self.add_visible_tiles(player_idx, end_loc[0], end_loc[1])

    def valid_shot(self, player: TrainPlayer, shot_row: int, shot_col: int) -> bool:
        if player is None:  # Player not in game
            return False

        if player.done:
            return False

        if player.shots:
            base_coords = tuple(player.shots[-1].location)
        else:
            if (shot_row, shot_col) == player.start:
                return True
            else:
                return False

        if shot_row > self.size[1] or shot_col > self.size[0] or shot_row < 1 or shot_col < 1:  # Out of bounds shots
            return False

        if len(self.board[(shot_row, shot_col)]["rails"]) >= 2:  # Tiles with too many players on them
            return False

        if abs(shot_row - base_coords[0]) == 1 and shot_col - base_coords[1] == 0:
            direction = "horizontal"
        elif abs(shot_col - base_coords[1]) == 1 and shot_row - base_coords[0] == 0:
            direction = "vertical"
        else:  # Shots not adjacent to player's rail endpoint
            return False
        base_intersecting_tag = None

        for tag in self.board[base_coords]["rails"]:  # Shots that move along someone else's rails for more than 1 tile
            if tag != player.tag:
                base_intersecting_tag = tag
        if base_intersecting_tag in self.board[(shot_row, shot_col)]:
            return False

        # Tiles that run next to your current rails
        if direction == "horizontal":
            test_cols = (shot_col - 1, shot_col + 1)
            for col in test_cols:
                if col > self.size[0] or col < 1:
                    pass
                elif player.tag in self.board[(shot_row, col)]["rails"]:
                    return False
        else:
            test_rows = (shot_row - 1, shot_row + 1)
            for row in test_rows:
                if row > self.size[1] or row < 1:
                    pass
                elif player.tag in self.board[(row, shot_col)]["rails"]:
                    return False
        return True

    async def push_player_update(self, ctx: interactions.SlashContext, p: TrainPlayer, p_idx: int):
        try:
            board_name = str(p.member.id)
            self.draw_board_img(
                filepath=f"{bd.parent}/Guilds/{ctx.guild_id}/Trains/{self.name}",
                board_name=str(p.member.id),
                player_idx=p_idx, player_board=True
            )
            board_img_path = f"{bd.parent}/Guilds/{ctx.guild_id}/Trains/{self.name}/{board_name}.png"
            await p.dmchannel.send(
                file=interactions.File(board_img_path),
                content=f"## Train board update for \"{self.name}\" in {ctx.guild.name}!"
            )
        except AttributeError:
            print(colored(f"Could not find user with ID {p.member.id}", "yellow"))
            del self.players[p_idx]

    async def update_boards_after_shot(
            self, ctx: interactions.SlashContext,
            row: int, column: int
    ) -> None:

        # Push updates to player boards, check if game is finished

        tasks = []
        for player_idx, player in enumerate(self.players):
            if (row, column) in player.vis_tiles:
                tasks.append(asyncio.create_task(self.push_player_update(ctx, player, player_idx)))

        await asyncio.gather(*tasks)
        active = not self.is_done()
        self.active = active
        if self.active:
            bd.active_trains[ctx.guild_id] = self
        else:
            del bd.active_trains[ctx.guild_id]

        await ctx.send(content=bd.pass_str, ephemeral=True)
        self.save_game(f"{bd.parent}/Guilds/{ctx.guild_id}/Trains/{self.name}")
        return None

    async def update_boards_after_create(
            self, ctx: interactions.SlashContext
    ) -> None:

        tasks = []

        for player_idx, player in enumerate(self.players):
            tasks.append(asyncio.create_task(self.push_player_update(ctx, player, player_idx)))
        await asyncio.gather(*tasks)

        # Update master board/game state
        self.save_game(f"{bd.parent}/Guilds/{ctx.guild_id}/Trains/{self.name}")
        bd.active_trains[ctx.guild_id] = self

        return None

    def gen_stats_embed(self, ctx: Union[interactions.SlashContext, interactions.ComponentContext],
                        page: int, expired: bool) -> tuple[interactions.Embed, Union[None, interactions.File]]:
        embed = interactions.Embed()
        embed.set_author(name=f"Anime Trains", icon_url=bd.bot_avatar_url)
        footer_end = ' | This message is inactive.' if expired else ' | This message deactivates after 5 minutes.'

        max_pages = len(self.players) + 1
        page = 1 + (page % max_pages)  # Loop back through pages both ways
        embed.set_footer(text=f'Page {page}/{max_pages} {footer_end}')

        # Game stats page
        if page == 1:
            resource_count = {}
            claimed_resource_count = {}
            rail_count = 0
            intersection_count = 0
            for coord, tile in self.board.items():
                if tile["resource"]:
                    try:
                        resource_count[tile["resource"]] += 1
                    except KeyError:
                        resource_count[tile["resource"]] = 1
                if tile["rails"]:
                    rail_count += 1
                    if tile["resource"]:
                        try:
                            claimed_resource_count[tile["resource"]] += 1
                        except KeyError:
                            claimed_resource_count[tile["resource"]] = 1
                    if len(tile["rails"]) > 1:
                        intersection_count += 1

            embed.title = "Game Stats"
            embed.description = f"*{self.name}*\n\u200b"
            embed.set_thumbnail(url=ctx.guild.icon.url)
            embed.add_field(name="🚂 Active?", value="✅" if self.active else "❌", inline=True)
            embed.add_field(name="🚂 Complete?", value="✅" if self.is_done() else "❌", inline=True)
            embed.add_field(name="\u200b", value="\u200b", inline=False)

            for resource, count in resource_count.items():
                if resource not in claimed_resource_count.keys():
                    claimed_resource_count[resource] = 0

                embed.add_field(
                    name=f"# of {resource} Claimed/Total",
                    value=f"{claimed_resource_count[resource]}/{count}", inline=True
                )

            embed.add_field(name="🛤️ Total Rails", value=rail_count, inline=True)
            embed.add_field(name="🔀 # of Crossings", value=intersection_count, inline=True)

            if self.is_done():
                self.draw_board_img(
                    filepath=f"{bd.parent}/Guilds/{ctx.guild_id}/Trains/{self.name}",
                    board_name="MASTER", player_board=False
                )
                board_img_path = f"{bd.parent}/Guilds/{ctx.guild_id}/Trains/{self.name}/MASTER.png"
                image = interactions.File(board_img_path, file_name="MASTER.png")
                embed.set_image(
                    url="attachment://MASTER.png"
                )
            else:
                with open(f"{bd.parent}/Data/nothing.png", "rb") as f:
                    file = io.BytesIO(f.read())
                image = interactions.File(file, file_name="there is nothing here")
            return embed, image

        # Player stats page
        player_idx = page - 2
        player = self.players[player_idx]

        if len(player.shots) == 0:
            embed.description = f"### {player.member.mention} has not placed any rails yet!"
            with open(f"{bd.parent}/Data/nothing.png", "rb") as f:
                file = io.BytesIO(f.read())
            image = interactions.File(file, file_name="there is nothing here")
            return embed, image

        embed.set_thumbnail(url=player.member.avatar_url)
        embed.description = f"### Stats for {player.member.mention}"
        embed.add_field(name="\u200b", value="\u200b", inline=False)

        # Total shots/in-zone shots
        total_shots = len(player.shots)
        in_zone_shots = 0
        shot_time = datetime.strptime(self.date, bd.date_format)
        time_between_shots_list = []
        for shot_idx, shot in enumerate(player.shots):
            if shot.genre == self.board[shot.location]["zone"]:
                in_zone_shots += 1

            time_between_shots_list.append((datetime.strptime(shot.time, bd.date_format) - shot_time).total_seconds())
            shot_time = datetime.strptime(shot.time, bd.date_format)

        avg_secs_between_shots = round(sum(time_between_shots_list)/len(time_between_shots_list))

        embed.add_field(name="🧮 Total Shots", value=total_shots, inline=True)
        embed.add_field(name="🍥 % in Zone", value=f"{round(in_zone_shots/total_shots*100)}%", inline=True)
        embed.add_field(name="🚂 Done?", value="✅" if player.done else "❌", inline=True)
        embed.add_field(
            name="⏳ Avg. Time Between Shots", value=str(timedelta(seconds=avg_secs_between_shots)), inline=False
        )

        # Projected completion time
        last_shot = player.shots[-1]
        dist_left = abs(last_shot.location[0] - player.end[0]) + abs(last_shot.location[1] - player.end[1])
        projected_time = datetime.now() + timedelta(seconds=round(dist_left*1.5)*avg_secs_between_shots)

        embed.add_field(
            name="🗓️ Projected Completion Date", value=projected_time.strftime("%Y/%m/%d at %H:%M:%S"), inline=False
        )

        # Shot genre pie chart

        genre_counts = self.players[player_idx].get_shot_genre_count()
        plt.style.use("dark_background")
        fig, ax = plt.subplots()

        plt.rcParams["font.size"] = 14
        plt.rcParams["font.family"] = "gg sans"
        plt.rcParams["font.weight"] = "bold"

        wedges, text, autotexts = ax.pie(genre_counts.values(), autopct="%1.1f%%")
        plt.setp(autotexts, size=16, weight="medium", color="black")
        plt.title(
            label="Shot Genre Percentages              ",
            weight="bold", size=17, family="gg sans", horizontalalignment="right"
        )
        plt.legend(
            genre_counts.keys(), title="Genres", loc="lower left", framealpha=0, bbox_to_anchor=(-0.45, 0.6, 0.75, 1),
            prop=matplotlib.font_manager.FontProperties(family="gg sans", weight="medium", size=15, style="italic"),
            title_fontproperties=matplotlib.font_manager.FontProperties(family="gg sans", weight="medium", size=17)
        )
        filepath = f"{bd.parent}/Guilds/{ctx.guild_id}/Trains/{self.name}/stats_img.png"
        plt.savefig(filepath, transparent=True)
        plt.close(fig)

        with open(filepath, "rb") as f:
            file = io.BytesIO(f.read())
        image = interactions.File(file, file_name="stats_img.png")

        embed.set_image(
            url="attachment://stats_img.png"
        )
        return embed, image

    def draw_board_img(
            self, filepath: str, board_name: str, player_board: bool = False, player_idx: int = 0,

    ):
        # Generate board image. Full board, if player board: only generate tiles which are rendered.
        # Grey out other tiles.

        label_offset = 1
        label_font_size = 24
        font = ImageFont.truetype(f"{bd.parent}/Data/ggsans/ggsans-Bold.ttf", label_font_size)
        tile_pixels = 50

        board_img = Image.new(
            mode="RGB",
            size=((self.size[0]+label_offset)*tile_pixels, (self.size[1]+label_offset)*tile_pixels),
            color=0xFFFFFF
        )
        draw = ImageDraw.Draw(board_img)
        hidden_tile_color = (255, 255, 255)
        border_color = (190, 190, 190)
        font_color = (0, 0, 0)

        def draw_hatch_pattern(row: int, col: int):
            hatch_color = (40, 40, 40)
            padding = 1

            x_start = tile_pixels*(col - 1)
            y_start = tile_pixels*(row - 1)

            x, y = x_start, y_start
            while y < y_start + tile_pixels:
                xy = ((x_start+padding, y+padding), (x+tile_pixels-padding, y_start+tile_pixels-padding))
                draw.line(xy=xy, fill=hatch_color, width=1)
                x -= 4
                y += 4

            x, y = x_start, y_start
            while x < x_start + tile_pixels:
                xy = ((x+padding, y_start+padding), (x_start+tile_pixels-padding, y+tile_pixels-padding))
                draw.line(xy=xy, fill=hatch_color, width=1)
                x += 4
                y -= 4

        # Draw column labels/tile borders
        label_padding = tile_pixels/2 - label_font_size/2
        for label_x in range(1, self.size[0]+1):

            draw.rectangle(
                xy=((label_x*tile_pixels, 1), ((label_x+1)*tile_pixels, tile_pixels)),
                fill=hidden_tile_color, outline=border_color, width=1
            )
            horizontal_offset = tile_pixels/2 - draw.textlength(text=str(label_x), font=font)/2
            draw.text(
                xy=(label_x*tile_pixels + horizontal_offset, label_padding), text=str(label_x), font=font, anchor="mm",
                fill=font_color
            )
        # Draw row labels/tile borders
        for label_y in range(1, self.size[1]+1):
            draw.rectangle(
                xy=((1, label_y*tile_pixels), (tile_pixels, (label_y+1)*tile_pixels)),
                fill=hidden_tile_color, outline=border_color, width=1
            )
            horizontal_offset = tile_pixels/2 - draw.textlength(text=str(label_y), font=font)/2
            draw.text(
                xy=(horizontal_offset, label_y*tile_pixels + label_padding), text=str(label_y), font=font, anchor="mm",
                fill=font_color
            )

        # Draw game tiles
        player = self.players[player_idx]

        for coords in self.board.keys():

            # Add label offset to leave space for row/col labels
            (row, col) = coords[0]+label_offset, coords[1]+label_offset

            # Draw hidden tile as gray, skip to next tile
            if player_board and coords not in player.vis_tiles:
                draw.rectangle(
                    xy=((col*tile_pixels, row*tile_pixels), (col*tile_pixels+tile_pixels, row*tile_pixels+tile_pixels)),
                    fill=hidden_tile_color, outline=border_color, width=1
                )
                continue

            # Draw non-hidden tiles

            tile_zone = self.board[coords]["zone"]
            if tile_zone is None:
                tile_color = (255, 255, 255)
            else:
                tile_color = genre_colors[tile_zone]

            draw.rectangle(
                xy=((col*tile_pixels, row*tile_pixels), (col*tile_pixels+tile_pixels, row*tile_pixels+tile_pixels)),
                fill=tile_color, outline=border_color, width=1
            )

            resource_text = demojize(self.board[coords]["resource"]) if self.board[coords]["resource"] else ""
            rail_text = "".join(self.board[coords]["rails"])

            if self.board[coords]["terrain"] == "river":
                draw_hatch_pattern(row, col)
            draw.text(xy=(col*tile_pixels, row*tile_pixels), text=rail_text+resource_text, anchor="mm", fill=font_color)

        try:
            board_img.save(f"{filepath}/{board_name}.png")
        except PermissionError:
            return "Could not write the board. Try again in a few seconds..."
        return None


# Function Definitions


# Manually call to make a zone image if needed
"""
def save_zones_img(filepath: str) -> None:
    wb = Workbook()
    ws = wb.active
    for idx, key in enumerate(genre_colors):
        row = idx // 4 + 1
        col = idx % 4 + 1
        ws.cell(row, col).value = key
        ws.cell(row, col).fill = PatternFill(patternType="solid", fgColor=genre_colors[key])
        ws.cell(row, col).font = Font(size=18, bold=True)
        ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 75
        ws.column_dimensions[get_column_letter(col)].width = 14

    wb.save(f"{filepath}/zones.xlsx")
    export_img(f"{filepath}/zones.xlsx", f"{filepath}/zones.png", "Sheet", "A1:D4")
    remove(f"{filepath}/zones.xlsx")
"""


async def load_game(
        filepath: str, bot: interactions.Client, guild: interactions.Guild, active_only: bool = False
) -> TrainGame:
    with open(f"{filepath}/gamedata.json", "r") as f:
        game_dict = json.load(f)

    if active_only and not game_dict["active"]:  # Skip loading inactive games if specified for faster loads
        return TrainGame(active=False)

    # Convert str/list keys back into tuple for use in game
    board = {}
    for key in game_dict["board"].keys():
        coords = key[1:-1].split(",")
        coords[0] = int(coords[0])
        coords[1] = int(coords[1])
        coords = tuple(coords)
        board[coords] = game_dict["board"][key]

    # Convert player dicts back into player classes
    player_list = []
    for player in game_dict["players"]:

        shot_list = []
        for shot in player["shots"]:
            shot_list.append(
                TrainShot(location=tuple(shot["location"]), genre=shot["genre"], info=shot["info"], time=shot["time"])
            )
        member = await guild.fetch_member(player["member_id"])
        player_list.append(
            TrainPlayer(
                member=member, tag=player["tag"], done=player["done"],
                rails=player["rails"], dmchannel=await bot.fetch_channel(player["dmchannel"]),
                start=tuple(player["start"]), end=tuple(player["end"]), score=player["score"], shots=shot_list,
                vis_tiles=[tuple(tile) for tile in player["vis_tiles"]], donetime=player["donetime"]
            )
        )

    game = TrainGame(
        name=game_dict["name"],
        date=game_dict["date"],
        players=player_list,
        board=board,
        gameid=game_dict["gameid"],
        active=game_dict["active"],
        size=tuple(game_dict["size"])
    )
    return game


def dict_to_rsp(rsp_dict: dict):
    if not rsp_dict:
        return None
    try:
        rsp = Response(rsp_dict['exact'], rsp_dict['trig'], rsp_dict['text'], rsp_dict['user_id'])
    except KeyError:
        if rsp_dict['m']:
            rsp_dict['exact'] = False
        else:
            rsp_dict['exact'] = True
        rsp_dict.pop('m')
        rsp = Response(rsp_dict['exact'], rsp_dict['trig'], rsp_dict['text'], rsp_dict['user_id'])
    return rsp


def dict_to_choices(dictionary: dict) -> list[interactions.SlashCommandChoice]:
    out = []
    for key in dictionary.keys():
        out.append(interactions.SlashCommandChoice(name=key, value=key))
    return out


def add_response(guild_id, rsp):
    f_name = 'responses.txt' if rsp.exact else 'mentions.txt'
    rsp.trig, rsp.text = demojize(rsp.trig), demojize(rsp.text)
    if not rsp.text:
        return True
    try:
        with open(f'{bd.parent}/Guilds/{guild_id}/{f_name}', 'r') as f:
            try:
                lines = json.load(f)
            except json.decoder.JSONDecodeError:
                lines = []
        duplicates = [next((x for x in lines if x['trig'] == rsp.trig), None)]
        if duplicates != [None]:
            for x in duplicates:
                if x['text'] == rsp.text:  # Reject identical additions
                    return True
        lines.append(rsp.asdict())
    except FileNotFoundError:
        f = open(f'{bd.parent}/Guilds/{guild_id}/{f_name}', 'w')
        f.close()
    try:
        with open(f'{bd.parent}/Guilds/{guild_id}/{f_name}', 'w') as f:
            json.dump(lines, f, indent=4)
    except UnicodeError:
        return True


def rmv_response(guild_id, rsp):
    f_name = 'responses.txt' if rsp.exact else 'mentions.txt'
    try:
        with open(f'{bd.parent}/Guilds/{guild_id}/{f_name}', 'r') as f:
            try:
                lines = json.load(f)
            except json.decoder.JSONDecodeError:
                return True
        to_del = [i for i, x in enumerate(lines) if x['trig'] == demojize(rsp.trig)]  # All matching entries
        if len(to_del) > 1:
            to_del = [
                i for i, x in enumerate(lines) if
                x['trig'] == demojize(rsp.trig) and x['text'].lower() == demojize(rsp.text.lower())
            ]
        if not to_del:
            return True
        k = 0
        for i in to_del:
            lines.pop(i - k)
            k += 1
        with open(f'{bd.parent}/Guilds/{guild_id}/{f_name}', 'w') as f:
            if not lines:
                f.write('[]')
            else:
                json.dump(lines, f, indent=4)

    except FileNotFoundError or ValueError:
        return True


def load_fonts(filepath):
    for font in matplotlib.font_manager.findSystemFonts(filepath):
        matplotlib.font_manager.fontManager.addfont(font)


def load_responses(file):
    lines = []
    try:
        with open(file, 'r') as f:
            try:
                lines = json.load(f)
            except ValueError:
                lines = []
        for idx, line in enumerate(lines):
            lines[idx] = dict_to_rsp(line)
        for rsp in lines:
            rsp.trig = emojize(rsp.trig)
            rsp.text = emojize(rsp.text)
    except FileNotFoundError:
        f = open(file, 'w')
        f.close()
    return lines


def get_resp(guild_id, trig, text, exact):
    if exact:
        for rsp in bd.responses[guild_id]:
            if rsp.trig == trig:
                if not text:
                    return rsp
                if rsp.text == text:
                    return rsp
    else:
        for rsp in bd.mentions[guild_id]:
            if rsp.trig == trig:
                if not text:
                    return rsp
                if rsp.text == text:
                    return rsp


def load_config(guild: interactions.Guild):
    # Load and validate guild bd.configs
    try:
        bd.config[int(guild.id)] = yaml.load(
            open(f'{bd.parent}/Guilds/{str(guild.id)}/config.yaml'),
            Loader=yaml.Loader
        )

        # Add missing keys
        for key in bd.default_config.keys():
            if key not in bd.config[int(guild.id)].keys():
                bd.config[int(guild.id)][key] = bd.default_config[key]
                print(
                    colored(f'{strftime("%Y-%m-%d %H:%M:%S")} :  ', 'white') +
                    colored(f'Config file for {guild.name} missing {key}, set to default.', 'yellow')
                )
                with open(f'{bd.parent}/Guilds/{int(guild.id)}/config.yaml', 'w') as f:
                    yaml.dump(bd.config[int(guild.id)], f, Dumper=yaml.Dumper)

        # Remove invalid keys
        temp = dict(bd.config[int(guild.id)])
        for key in bd.config[int(guild.id)].keys():
            if key not in bd.default_config.keys():
                temp = dict(bd.config[int(guild.id)])
                del temp[key]
                print(
                    colored(f'{strftime("%Y-%m-%d %H:%M:%S")} :  ', 'white') +
                    colored(f'Invalid key {key} in {guild.name} config, removed.', 'yellow')
                )
                with open(f'{bd.parent}/Guilds/{int(guild.id)}/config.yaml', 'w') as f:
                    yaml.dump(temp, f, Dumper=yaml.Dumper)
        bd.config[int(guild.id)] = temp

    # Create new file if config is missing
    except FileNotFoundError:
        with open(f'{bd.parent}/Guilds/{int(guild.id)}/config.yaml', 'w') as f:
            yaml.dump(bd.default_config, f, Dumper=yaml.Dumper)
        print(
            colored(f'{strftime("%Y-%m-%d %H:%M:%S")} :  ', 'white') +
            colored(f'No config file found for {guild.name}, created default config file.', 'yellow')
        )
        bd.config[int(guild.id)] = yaml.load(open(f'{bd.parent}/Guilds/{guild.id}/config.yaml'), Loader=yaml.Loader)


def gen_resp_list(guild: interactions.Guild, page: int, expired: bool) -> interactions.Embed:

    guild_id = int(guild.id)
    list_msg = interactions.Embed(
        description='*Your response list, sir.*'
    )
    guild_trigs = []
    for rsp in bd.responses[guild_id]:
        guild_trigs.append(rsp)
    for mtn in bd.mentions[guild_id]:
        guild_trigs.append(mtn)

    max_pages = 1 if len(guild_trigs) <= 10 else len(guild_trigs) // 10 + 1  # Determine max pg at 10 entries per pg
    page = 1 + (page % max_pages)  # Loop back through pages both ways
    footer_end = ' | This message is inactive.' if expired else ' | This message deactivates after 5 minutes.'
    list_msg.set_author(name=guild.name, icon_url=bd.bot_avatar_url)
    list_msg.set_thumbnail(url=guild.icon.url)
    list_msg.set_footer(text=f'Page {page}/{max_pages} {footer_end}')
    nums = range((page-1)*10, len(guild_trigs)) if page == max_pages else range((page-1)*10, page*10)
    for i in nums:
        pref = '**Exact Trigger:** ' if guild_trigs[i].exact else '**Phrase Trigger:** '
        rsp_field = f'{pref}{guild_trigs[i].trig} \n **Respond:** {guild_trigs[i].text}'
        if len(rsp_field) >= 1024:
            rsp_field = f'{pref}{guild_trigs[i].trig} \n **Respond:** *[Really, really, really long response]*'

        list_msg.add_field(
            name='\u200b',
            value=rsp_field, inline=False
        )
    return list_msg


async def close_msg(list_msg: ListMsg, delay: int, ctx: interactions.SlashContext, msg: interactions.Message) -> None:
    await asyncio.sleep(delay)

    if list_msg.msg_type == "rsplist":
        embed = gen_resp_list(ctx.guild, list_msg.page, True)
    elif list_msg.msg_type == "trainrules":
        embed = gen_rules_embed(list_msg.page, True)
    else:
        embed = None
    await msg.edit(embeds=embed)
    bd.active_msgs.remove(list_msg)


async def get_members_from_str(guild, txt: str) -> list[interactions.Member]:
    mentions = []
    mention = ""
    mention_start = False
    id_start = False

    for char in txt:

        # Begin recording of characters until a non-integer character is encountered
        if id_start:
            try:
                int(char)
                mention += char
            except ValueError:
                try:
                    if int(mention) in mentions:
                        continue
                    mentions.append(int(mention))
                except ValueError:
                    pass
                id_start = False
                mention_start = False
                mention = ""

        # Confirm start of discord mention string
        if char == "@" and mention_start:
            id_start = True

        # Mark start of discord mention string
        if char == "<":
            mention_start = True

    # Check for invalid player IDs
    members = []
    for entry in mentions:
        member = await guild.fetch_member(entry)
        if not member or member.bot:
            pass
        else:
            members.append(member)
    return members


def get_player_tags(users: list[interactions.Member]) -> list[str]:
    tags = []
    for user in users:
        done = False
        for idx, letter in enumerate(user.global_name):
            if user.global_name[0:idx+1] not in tags:
                tags.append(user.global_name[0:idx+1].upper())
                done = True
                break
        if not done:
            tags.append(user.global_name.upper())
    return tags


def gen_rules_embed(page: int, expired: bool) -> interactions.Embed:
    max_pages = 5
    page = 1 + (page % max_pages)  # Loop back through pages both ways
    footer_end = ' | This message is inactive.' if expired else ' | This message deactivates after 5 minutes.'
    if page == 1:
        embed = train_rules_embed()
    elif page == 2:
        embed = train_zones_embed()
    elif page == 3:
        embed = train_symbols_embed()
    elif page == 4:
        embed = train_quests_embed()
    else:
        embed = train_scoring_embed()
    embed.set_footer(text=f'Page {page}/{max_pages} {footer_end}')
    return embed


# Embed/Component Definitions

"""
if "rempowerbed" in message.content:
    await channel.send(
        content='Wished by <@493498108211232779>',
        embeds=bu.power_embed(),
        components=bu.wish_button()
    )
    await channel.send(
        content='Wished by <@302266697488924672>',
        embeds=bu.rembed(),
        components=bu.wish_button()
    )
        
def rembed():
    embed = interactions.Embed(
    )
    embed.color = 0xff9c2c
    embed.add_field(
        name='Rem',
        inline=False,
        value='\u200b\nRe:Zero kara Hajimeru Isekai Seikatsu\n'
              '**1528** <:Kakera:1103395628723228732>\n'
              'React with any emoji to claim!',
    )
    embed.set_image(url='https://mudae.net/uploads/4190198/-00NfGVxGabXCbZZfHPc~bFtvJih.png')
    return embed


def power_embed():
    embed = interactions.Embed(
    )
    embed.color = 0xff9c2c
    embed.add_field(
        name='Power',
        inline=False,
        value='\u200b\nChainsaw Man\n'
              '**1364** <:Kakera:1103395628723228732>\n'
              'React with any emoji to claim!',
    )
    embed.set_image(url='https://mudae.net/uploads/7637289/JW6Pl0JPh04sSEnVARM3~qcyFORS.png')
    return embed
"""


def train_rules_embed():
    embed = interactions.Embed()
    embed.set_author(name="Anime Trains", icon_url=bd.bot_avatar_url)
    embed.color = 0xff9c2c
    embed.title = "Rules"
    embed.description = \
        "**1.** Each shot (3 hours) corresponds to one track being placed down.\n\n" \
        "**2.** Rails may intersect. Each intersection awards both players with an extra point. " \
        "No more than two player's rails can intersect in the same location.\n\n" \
        "**3.** Rails may be placed directly adjacent to any existing rails, but it must follow a single path. " \
        "(No self-intersections). In addition, there must be at least one space between your rails " \
        "unless they are connected.\n\n" \
        "**4.** The board is generated randomly at the start of the game, and the points are tallied at " \
        "the end of the game. The player with the most points at the end of the game wins.\n\n"
    return embed


def train_zones_embed():
    embed = interactions.Embed()
    embed.set_author(name="Anime Trains", icon_url=bd.bot_avatar_url)
    embed.color = 0xff9c2c
    embed.title = "Genre Zones"
    embed.description = "If the primary show used for a shot has a genre matching the genre zone of the shot, " \
                        "only 1/2 of the usual amount of rails are consumed."

    embed.add_field(
        name="\u200b",
        inline=False,
        value="**Genre zones appear as the following colors on the trains board:**"
    )
    embed.set_image(url=bd.train_zones_url)
    return embed


def train_symbols_embed():
    embed = interactions.Embed()
    embed.set_author(name="Anime Trains", icon_url=bd.bot_avatar_url)
    embed.color = 0xff9c2c
    embed.title = "Symbol Reference"
    embed.add_field(
        name="🌾: Wheat",
        value="Plus 1 point if connected to your network. Plus 2 more points if connected to a city. "
              "Each additional wheat is worth 1 point only.",
        inline=True
    )
    embed.add_field(
        name="🌳: Wood",
        value="Provides 2 points for each wood connected to your network.",
        inline=True
    )
    embed.add_field(
        name="💎: Gems",
        value="Provides 2 points if connected to your network. "
              "The first player to connect gems to their network gets 3 bonus points.",
        inline=True
    )
    embed.add_field(
        name="\u200b",
        value="\u200b",
        inline=False
    )
    embed.add_field(
        name="🌃: City",
        value="Provides above bonuses. Each city has a favorite season (revealed at end). "
              "Any player who shoots a city with the correct season gets 3 bonus points.",
        inline=True
    )
    embed.add_field(
        name="⛓: Prison",
        value="Must be shot with a show with one of the following tags: Feet, Loli (Character), Maid, Nudity. "
              "Once the prisoners are released, everybody else loses 2 points.",
        inline=True
    )
    embed.add_field(
        name="🏠: House",
        value="Provides 1 points for each house connected to your network. "
              "If the house is connected to a city, then the player gains 2 bonus points. "
              "If the house is connected to a prison, the player loses 1 point.",
        inline=True
    )
    embed.add_field(
        name="🏞️ Gray dotted tiles: River",
        value="Shots made on rivers use double the normal amount of rails.",
        inline=False
    )
    return embed


def train_quests_embed():
    embed = interactions.Embed()
    embed.set_author(name="Anime Trains", icon_url=bd.bot_avatar_url)
    embed.color = 0xff9c2c
    embed.title = "Quests"
    embed.description = "*Quests may be completed by every player once.*"

    embed.add_field(
        name="\u200b",
        value="**1.** Completely watch one show with the \"trains\" tag. **Reward: 3**\n\n"
              "**2.** Make at least two shots of your least watched genre (excluding Hentai). **Reward: 4**\n\n"
              "**3.** Make a shot of shows with each of the following sources: "
              "Anime original, manga, light novel, mugi original. **Reward: 3**\n\n"
              "**4.** Do not make a single shot of a genre on its corresponding zone. **Reward: 3**\n\n"
              "**5.** Make shots with at least three shows from another player's list. **Reward: 2**",
        inline=False
    )
    return embed


def train_scoring_embed():
    embed = interactions.Embed()
    embed.set_author(name="Anime Trains", icon_url=bd.bot_avatar_url)
    embed.color = 0xff9c2c
    embed.title = "Scoring"
    embed.description = \
        "**1.** Each player's score is calculated at the end of the game.\n\n" \
        "**2.** Players earn points from the sources listed below:\n" \
        "- The 1st/2nd player to finish their track earn 2/1 bonus points respectively.\n" \
        "- Points from resources (see page 3) and points from quests (see page 4).\n" \
        "- Players earn 1 point each time their track intersects another player's track.\n" \
        "- For every 2 rails less than 30 that a player uses, they gain a point. " \
        "For every two rails over 30, that player loses a point.\n\n" \
        "3. Automatic score calculation (minus quests) is a planned feature!"
    return embed


def train_game_embed(ctx: interactions.SlashContext, game: TrainGame):
    embed = interactions.Embed()
    embed.set_author(name="Anime Trains", icon_url=bd.bot_avatar_url)
    embed.color = 0xff9c2c
    embed.title = "It's Train Time"
    embed.description = f"*{ctx.author.mention} has created \"{game.name}\"!*"
    embed.set_thumbnail(url=ctx.author.avatar_url)

    embed.add_field(
        name="Board Size",
        value=f"{game.size[0]} by {game.size[1]}",
        inline=True
    )
    player_mentions = []
    for player in game.players:
        player_mentions.append(f"<@{player.member.id}>")
    embed.add_field(
        name="Players",
        value=", ".join(player_mentions),
        inline=True
    )
    embed.add_field(
        name="\u200b",
        value="**Players, check your DMs to see your board!**"
    )
    embed.set_footer(text=game.date)

    return embed


def wish_button():
    return interactions.Button(
        style=interactions.ButtonStyle.SECONDARY,
        label='💕',
        custom_id='wish'
    )


def prevpg_rsp():
    return interactions.Button(
        style=interactions.ButtonStyle.SECONDARY,
        label="←",
        custom_id="prevpg_rsp",
    )


def nextpg_rsp():
    return interactions.Button(
        style=interactions.ButtonStyle.SECONDARY,
        label="→️",
        custom_id="nextpg_rsp",
    )


def prevpg_trainrules():
    return interactions.Button(
        style=interactions.ButtonStyle.SECONDARY,
        label="←",
        custom_id="prevpg_trainrules",
    )


def nextpg_trainrules():
    return interactions.Button(
        style=interactions.ButtonStyle.SECONDARY,
        label="→️",
        custom_id="nextpg_trainrules",
    )


def prevpg_trainstats():
    return interactions.Button(
        style=interactions.ButtonStyle.SECONDARY,
        label="←",
        custom_id="prevpg_trainstats",
    )


def nextpg_trainstats():
    return interactions.Button(
        style=interactions.ButtonStyle.SECONDARY,
        label="→️",
        custom_id="nextpg_trainstats",
    )


genre_colors = {
    'Action': (255, 125, 125),
    'Adventure': (102, 255, 153),
    # 'Comedy': (136, 255, 136)),
    'Drama': (245, 197, 255),
    'Ecchi': (255, 204, 204),
    'Fantasy': (155, 194, 230),
    # 'Hentai': (255, 0, 255),
    'Horror': (169, 208, 142),
    # 'Mahou_Shoujo': (255, 255, 136),
    'Mecha': (217, 217, 217),
    'Music': (221, 235, 247),
    'Mystery': (174, 170, 170),
    'Psychological': (255, 217, 102),
    'Romance': (208, 125, 163),
    'Sci-Fi': (255, 242, 204),
    'SoL': (142, 169, 219),
    'Sports': (237, 125, 49),
    'Supernatural': (244, 176, 132),
    'Thriller': (226, 239, 218),
}
